import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import os
from datetime import datetime, timedelta
from collections import defaultdict
import math
import threading
import configparser

# PDF libraries
try:
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("ReportLab not installed. PDF export will be disabled.")
    print("Install with: pip install reportlab")

try:
    from tkcalendar import Calendar
except ImportError:
    Calendar = None  # Будет работать без календаря


def load_input_directory():
    """Загрузить путь к директории convert_data из config.ini"""
    config = configparser.ConfigParser()
    config_file = 'config.ini'
    
    if os.path.exists(config_file):
        config.read(config_file)
        # Читаем convert_data из config.ini
        convert_path = config.get('PATHS', 'convert_data', fallback=r'C:\Users\dotignore\Documents\Python\examplaone_krakenSDR_web\data')
        # Преобразуем '/' на '\' для Windows
        convert_path = convert_path.replace('/', '\\')
        return convert_path
    else:
        # Если config.ini не найден, используем путь по умолчанию
        return r'C:\Users\dotignore\Documents\Python\examplaone_krakenSDR_web\data'


class DateMaskEntry(tk.Frame):
    """Widget for date input with mask DD/MM/YY HH:MM:SS and calendar"""
    def __init__(self, parent, textvariable=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.textvariable = textvariable
        
        # Frame for entry and button
        input_frame = tk.Frame(self)
        input_frame.pack(side=tk.LEFT)
        
        # Create Entry
        self.entry = tk.Entry(input_frame, width=17, font=("Courier", 9))
        self.entry.pack(side=tk.LEFT)
        
        # Calendar icon button
        self.cal_button = tk.Button(input_frame, text="📅", width=2, 
                                   command=self.show_calendar,
                                   font=("Arial", 10), cursor="hand2",
                                   relief=tk.FLAT, bg="#f5f5f5")
        self.cal_button.pack(side=tk.LEFT, padx=(2, 0))
        
        # Initial value
        self.entry.insert(0, "01/01/25 00:00:00")
        
        # Bind events
        self.entry.bind('<KeyPress>', self.on_key_press)
        self.entry.bind('<Left>', self.on_arrow_left)
        self.entry.bind('<Right>', self.on_arrow_right)
        
        # Separator positions
        self.separators = {2: '/', 5: '/', 8: ' ', 11: ':', 14: ':'}
        self.field_positions = [
            (0, 2),   # DD
            (3, 5),   # MM
            (6, 8),   # YY
            (9, 11),  # HH
            (12, 14), # MM
            (15, 17)  # SS
        ]
        
        # Synchronization with textvariable
        if self.textvariable:
            self.textvariable.trace('w', self.update_from_var)
            self.update_from_var()
    
    def on_arrow_left(self, event):
        """Handle left arrow key"""
        pos = self.entry.index(tk.INSERT)
        
        # Skip separators when moving left
        if pos > 0:
            pos -= 1
            if pos in self.separators:
                pos -= 1
            self.entry.icursor(pos)
        return 'break'
    
    def on_arrow_right(self, event):
        """Handle right arrow key"""
        pos = self.entry.index(tk.INSERT)
        
        # Skip separators when moving right
        if pos < 16:
            pos += 1
            if pos in self.separators:
                pos += 1
            self.entry.icursor(pos)
        return 'break'
    
    def on_key_press(self, event):
        """Handle key press events"""
        if event.char.isdigit():
            pos = self.entry.index(tk.INSERT)
            
            # Don't allow input at separator positions
            if pos in self.separators:
                return 'break'
            
            # Replace character at cursor position
            text = list(self.entry.get())
            if pos < len(text):
                text[pos] = event.char
                self.entry.delete(0, tk.END)
                self.entry.insert(0, ''.join(text))
                
                # Don't move cursor automatically, keep in place
                new_pos = pos + 1
                
                # Skip separator only if next position is a separator
                if new_pos in self.separators:
                    new_pos += 1
                    
                # Check if we're within bounds
                if new_pos <= 16:
                    self.entry.icursor(new_pos)
                else:
                    self.entry.icursor(pos)
            
            # Update variable
            if self.textvariable:
                self.textvariable.set(self.entry.get())
                
            return 'break'
        
        elif event.keysym == 'BackSpace':
            pos = self.entry.index(tk.INSERT)
            if pos > 0 and pos - 1 not in self.separators:
                text = list(self.entry.get())
                text[pos - 1] = '0'
                self.entry.delete(0, tk.END)
                self.entry.insert(0, ''.join(text))
                self.entry.icursor(pos - 1)
            return 'break'
        
        elif event.keysym in ['Tab', 'Return']:
            return  # Allow standard behavior
        
        return 'break'
    
    def show_calendar(self):
        """Show calendar for date and time selection"""
        if not Calendar:
            print("tkcalendar not installed! Install: pip install tkcalendar")
            return
            
        cal_window = tk.Toplevel(self)
        cal_window.title("Select date and time")
        cal_window.geometry("320x520")
        cal_window.resizable(False, False)
        
        try:
            current_date_str = self.entry.get()
            current_date = datetime.strptime(current_date_str, "%d/%m/%y %H:%M:%S")
        except:
            current_date = datetime.now()
        
        # Main container with scrolling if needed
        main_container = tk.Frame(cal_window)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Frame for calendar
        date_frame = tk.LabelFrame(main_container, text="Date", padx=10, pady=5)
        date_frame.pack(fill=tk.X, padx=10, pady=5)
        
        cal = Calendar(date_frame, selectmode='day',
                    year=current_date.year, 
                    month=current_date.month,
                    day=current_date.day,
                    date_pattern='dd/mm/yyyy')
        cal.pack()
        
        # Frame for time
        time_frame = tk.LabelFrame(main_container, text="Time (click to select)", padx=10, pady=5)
        time_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Variables for selected time
        selected_hour = tk.IntVar(value=current_date.hour)
        selected_min = tk.IntVar(value=current_date.minute)
        selected_sec = tk.IntVar(value=current_date.second)
        
        # Frame for displaying selected time
        selected_time_frame = tk.Frame(time_frame)
        selected_time_frame.pack(pady=5)
        
        hour_label = tk.Label(selected_time_frame, text=f"{selected_hour.get():02d}", 
                            font=("Courier", 16, "bold"), width=2, relief=tk.SUNKEN, bg="white")
        hour_label.grid(row=0, column=0, padx=2)
        
        tk.Label(selected_time_frame, text=":", font=("Courier", 16, "bold")).grid(row=0, column=1)
        
        min_label = tk.Label(selected_time_frame, text=f"{selected_min.get():02d}", 
                            font=("Courier", 16, "bold"), width=2, relief=tk.SUNKEN, bg="white")
        min_label.grid(row=0, column=2, padx=2)
        
        tk.Label(selected_time_frame, text=":", font=("Courier", 16, "bold")).grid(row=0, column=3)
        
        sec_label = tk.Label(selected_time_frame, text=f"{selected_sec.get():02d}", 
                            font=("Courier", 16, "bold"), width=2, relief=tk.SUNKEN, bg="white")
        sec_label.grid(row=0, column=4, padx=2)
        
        # Frame with three columns for time selection
        picker_frame = tk.Frame(time_frame)
        picker_frame.pack(fill=tk.X)
        
        # Hours column
        hour_frame = tk.Frame(picker_frame)
        hour_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Label(hour_frame, text="Hours", font=("Arial", 9)).pack()
        
        hour_listbox = tk.Listbox(hour_frame, height=4, width=4, font=("Courier", 10))
        hour_scroll = tk.Scrollbar(hour_frame, orient="vertical", command=hour_listbox.yview)
        hour_listbox.configure(yscrollcommand=hour_scroll.set)
        
        for i in range(24):
            hour_listbox.insert(tk.END, f"{i:02d}")
        
        hour_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        hour_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        hour_listbox.selection_set(current_date.hour)
        hour_listbox.see(current_date.hour)
        
        # Minutes column
        min_frame = tk.Frame(picker_frame)
        min_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Label(min_frame, text="Minutes", font=("Arial", 9)).pack()
        
        min_listbox = tk.Listbox(min_frame, height=4, width=4, font=("Courier", 10))
        min_scroll = tk.Scrollbar(min_frame, orient="vertical", command=min_listbox.yview)
        min_listbox.configure(yscrollcommand=min_scroll.set)
        
        for i in range(60):
            min_listbox.insert(tk.END, f"{i:02d}")
        
        min_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        min_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        min_listbox.selection_set(current_date.minute)
        min_listbox.see(current_date.minute)
        
        # Seconds column
        sec_frame = tk.Frame(picker_frame)
        sec_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        tk.Label(sec_frame, text="Seconds", font=("Arial", 9)).pack()
        
        sec_listbox = tk.Listbox(sec_frame, height=4, width=4, font=("Courier", 10))
        sec_scroll = tk.Scrollbar(sec_frame, orient="vertical", command=sec_listbox.yview)
        sec_listbox.configure(yscrollcommand=sec_scroll.set)
        
        for i in range(60):
            sec_listbox.insert(tk.END, f"{i:02d}")
        
        sec_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        sec_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        sec_listbox.selection_set(current_date.second)
        sec_listbox.see(current_date.second)
        
        # Update functions on selection
        def on_hour_select(event):
            selection = hour_listbox.curselection()
            if selection:
                selected_hour.set(selection[0])
                hour_label.config(text=f"{selection[0]:02d}")
        
        def on_min_select(event):
            selection = min_listbox.curselection()
            if selection:
                selected_min.set(selection[0])
                min_label.config(text=f"{selection[0]:02d}")
        
        def on_sec_select(event):
            selection = sec_listbox.curselection()
            if selection:
                selected_sec.set(selection[0])
                sec_label.config(text=f"{selection[0]:02d}")
        
        hour_listbox.bind('<<ListboxSelect>>', on_hour_select)
        min_listbox.bind('<<ListboxSelect>>', on_min_select)
        sec_listbox.bind('<<ListboxSelect>>', on_sec_select)
        
        def apply_datetime():
            date_str = cal.get_date()
            selected_date = datetime.strptime(date_str, '%d/%m/%Y')
            
            final_date = selected_date.replace(hour=selected_hour.get(), 
                                            minute=selected_min.get(), 
                                            second=selected_sec.get())
            date_formatted = final_date.strftime("%d/%m/%y %H:%M:%S")
            
            self.entry.delete(0, tk.END)
            self.entry.insert(0, date_formatted)
            
            if self.textvariable:
                self.textvariable.set(date_formatted)
            
            cal_window.destroy()
        
        # Control buttons
        button_frame = tk.Frame(cal_window, bg="#e0e0e0", height=60)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X)
        button_frame.pack_propagate(False)
        
        ok_button = tk.Button(button_frame, 
                            text="✓ APPLY", 
                            command=apply_datetime,
                            bg="#4CAF50", 
                            fg="white", 
                            font=("Arial", 11, "bold"),
                            width=12,
                            height=2,
                            cursor="hand2")
        ok_button.pack(side=tk.LEFT, padx=20, pady=10)
        
        cancel_button = tk.Button(button_frame, 
                                text="✗ CANCEL", 
                                command=cal_window.destroy,
                                bg="#f44336", 
                                fg="white", 
                                font=("Arial", 11, "bold"),
                                width=12,
                                height=2,
                                cursor="hand2")
        cancel_button.pack(side=tk.RIGHT, padx=20, pady=10)
        
        # Make window modal
        cal_window.transient(self.winfo_toplevel())
        cal_window.grab_set()
        
        # Center window
        cal_window.update_idletasks()
        x = (cal_window.winfo_screenwidth() // 2) - 160
        y = (cal_window.winfo_screenheight() // 2) - 260
        cal_window.geometry(f"+{x}+{y}")

    def update_from_var(self, *args):
        """Обновление виджета из переменной"""
        value = self.textvariable.get()
        if value and len(value) == 17:
            current = self.entry.get()
            if current != value:
                self.entry.delete(0, tk.END)
                self.entry.insert(0, value)
    
    def get(self):
        """Получить значение"""
        return self.entry.get()


class ProgressWindow:
    """Window for showing export progress"""
    def __init__(self, parent, title="Export Progress"):
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("400x200")
        self.window.resizable(False, False)
        
        # Center window
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - 200
        y = (self.window.winfo_screenheight() // 2) - 100
        self.window.geometry(f"+{x}+{y}")
        
        # Make window modal
        self.window.transient(parent)
        self.window.grab_set()
        
        # Variables
        self.cancelled = False
        
        # Create UI
        self.setup_ui()
    
    def setup_ui(self):
        """Create progress window UI"""
        # Main frame
        main_frame = tk.Frame(self.window, bg="#f5f5f5")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_frame, text="Exporting to Excel", 
                              font=("Arial", 14, "bold"), bg="#f5f5f5")
        title_label.pack(pady=(0, 20))
        
        # Current file label
        self.current_file_label = tk.Label(main_frame, text="Processing: ...", 
                                          font=("Arial", 10), bg="#f5f5f5", 
                                          fg="#333333")
        self.current_file_label.pack(pady=(0, 10))
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                           maximum=100, length=300)
        self.progress_bar.pack(pady=(0, 10))
        
        # Progress text
        self.progress_text = tk.Label(main_frame, text="0%", 
                                     font=("Arial", 10), bg="#f5f5f5")
        self.progress_text.pack(pady=(0, 10))
        
        # Files processed
        self.files_label = tk.Label(main_frame, text="Files: 0/0 processed", 
                                   font=("Arial", 9), bg="#f5f5f5", 
                                   fg="#666666")
        self.files_label.pack(pady=(0, 20))
        
        # Cancel button
        self.cancel_button = tk.Button(main_frame, text="Cancel", 
                                      command=self.cancel_export,
                                      bg="#f44336", fg="white",
                                      font=("Arial", 10, "bold"),
                                      width=15, height=2)
        self.cancel_button.pack()
    
    def update_progress(self, current_file, progress, files_processed, total_files):
        """Update progress display"""
        self.current_file_label.config(text=f"Processing: {current_file}")
        self.progress_var.set(progress)
        self.progress_text.config(text=f"{progress:.0f}%")
        self.files_label.config(text=f"Files: {files_processed}/{total_files} processed")
        self.window.update()
    
    def cancel_export(self):
        """Cancel the export operation"""
        self.cancelled = True
        self.current_file_label.config(text="Cancelling...", fg="#f44336")
        self.cancel_button.config(text="Cancelling...", state="disabled")
    
    def close(self):
        """Close the progress window"""
        self.window.destroy()


class DMRDataViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("DMR Data Viewer")
        self.root.geometry("1600x900")
        
        # Input directory
        self.input_dir = load_input_directory()
        
        # Data storage
        self.file_data = []
        self.filtered_file_data = []
        
        # File selection
        self.selected_files = set()  # Set of selected file indices
        
        # Last save directory
        self.last_save_dir = os.path.expanduser("~")  # Default to user home directory
        
        # Sort variables for file selector
        self.sort_by_name = tk.StringVar(value="asc")  # asc, desc
        self.sort_by_sessions = tk.StringVar(value="none")  # asc, desc, none
        
        # Collapsed/expanded state for groups
        # By default, all groups are collapsed (not in the set means collapsed)
        self.expanded_groups = set()  # Set of (date_str, hour, filename, group_index) tuples
        
        # Filter variables
        self.date_from_var = tk.StringVar(value="01/01/25 00:00:00")
        self.date_to_var = tk.StringVar(value="31/12/25 23:59:59")
        self.duration_from_var = tk.StringVar(value="")
        self.duration_to_var = tk.StringVar(value="")
        self.selected_event = tk.StringVar(value="All")
        self.selected_timeslot = tk.StringVar(value="All")
        self.selected_color_code = tk.StringVar(value="All")
        self.selected_algorithm = tk.StringVar(value="All")
        self.selected_key = tk.StringVar(value="All")
        
        # Session grouping gap
        self.session_gap_var = tk.StringVar(value="15")
        
        # Unique values for filters
        self.unique_events = set()
        self.unique_timeslots = set()
        self.unique_color_codes = set()
        self.unique_algorithms = set()
        self.unique_keys = set()
        
        # Create interface
        self.setup_ui()
        
        # Load data
        self.load_all_data()
        self.set_date_range_from_data()
        self.update_comboboxes()
        self.display_data()
    
    def setup_ui(self):
        """Create user interface"""
        main_frame = tk.Frame(self.root, bg="#ffffff")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Top panel with buttons
        top_frame = tk.Frame(main_frame, bg="#ffffff", height=40)
        top_frame.pack(fill=tk.X)
        top_frame.pack_propagate(False)
        
        # Button panel
        button_panel = tk.Frame(top_frame, bg="#ffffff")
        button_panel.pack(side=tk.RIGHT, padx=10, pady=5)
        
        # Export to PDF button
        export_pdf_btn = tk.Button(button_panel, text="📄 Export PDF", 
                                 command=self.export_to_pdf,
                                 bg="#FF5722", fg="white",
                                 font=("Arial", 9, "bold"),
                                 relief=tk.RAISED, bd=1)
        export_pdf_btn.pack(side=tk.LEFT, padx=5)
        
        # Export to text files button
        export_txt_btn = tk.Button(button_panel, text="📝 Export TXT", 
                                 command=self.export_to_text_files,
                                 bg="#4CAF50", fg="white",
                                 font=("Arial", 9, "bold"),
                                 relief=tk.RAISED, bd=1)
        export_txt_btn.pack(side=tk.LEFT, padx=5)
        
        # Export to Excel button
        export_excel_btn = tk.Button(button_panel, text="📊 Export Excel", 
                                    command=self.export_to_excel,
                                    bg="#2E7D32", fg="white",
                                    font=("Arial", 9, "bold"),
                                    relief=tk.RAISED, bd=1)
        export_excel_btn.pack(side=tk.LEFT, padx=5)
        
        refresh_btn = tk.Button(button_panel, text="Refresh", command=self.refresh_all,
                              bg="#e0e0e0", relief=tk.RAISED, bd=1)
        refresh_btn.pack(side=tk.LEFT, padx=5)
        
        # Horizontal container
        content_frame = tk.Frame(main_frame, bg="#ffffff")
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Left filter panel
        filter_frame = tk.Frame(content_frame, bg="#f5f5f5", width=200)
        filter_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(5, 0))
        filter_frame.pack_propagate(False)
        
        # Container for filters
        filter_content = tk.Frame(filter_frame, bg="#f5f5f5")
        filter_content.pack(fill=tk.BOTH, expand=True, padx=10)
        
        # FILE SELECTION BUTTON
        file_select_frame = tk.Frame(filter_content, bg="#f5f5f5")
        file_select_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.select_files_button = tk.Button(file_select_frame, text="SELECT FILES", 
                                           command=self.open_file_selector,
                                           bg="#2196F3", fg="white",
                                           font=("Arial", 10, "bold"),
                                           relief=tk.RAISED, bd=2,
                                           cursor="hand2", width=15)
        self.select_files_button.pack(fill=tk.X)
        
        
        # DURATION FILTER
        duration_frame = tk.Frame(filter_content, bg="#f5f5f5")
        duration_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(duration_frame, text="DURATION", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        
        dur_from_frame = tk.Frame(duration_frame, bg="#f5f5f5")
        dur_from_frame.pack(fill=tk.X, pady=2)
        tk.Label(dur_from_frame, text="From:", bg="#f5f5f5", width=7).pack(side=tk.LEFT)
        tk.Entry(dur_from_frame, textvariable=self.duration_from_var, width=10).pack(side=tk.LEFT)
        tk.Label(dur_from_frame, text="sec", bg="#f5f5f5", font=("Arial", 9)).pack(side=tk.LEFT, padx=(2, 0))
        
        dur_to_frame = tk.Frame(duration_frame, bg="#f5f5f5")
        dur_to_frame.pack(fill=tk.X, pady=2)
        tk.Label(dur_to_frame, text="To:", bg="#f5f5f5", width=7).pack(side=tk.LEFT)
        tk.Entry(dur_to_frame, textvariable=self.duration_to_var, width=10).pack(side=tk.LEFT)
        tk.Label(dur_to_frame, text="sec", bg="#f5f5f5", font=("Arial", 9)).pack(side=tk.LEFT, padx=(2, 0))
        
        # EVENT FILTER
        event_frame = tk.Frame(filter_content, bg="#f5f5f5")
        event_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(event_frame, text="EVENT", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        self.event_combo = ttk.Combobox(event_frame, textvariable=self.selected_event, 
                                      state="readonly", width=20)
        self.event_combo.pack()
        
        # TIMESLOT FILTER
        timeslot_frame = tk.Frame(filter_content, bg="#f5f5f5")
        timeslot_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(timeslot_frame, text="TIMESLOT", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        self.timeslot_combo = ttk.Combobox(timeslot_frame, textvariable=self.selected_timeslot, 
                                         state="readonly", width=20)
        self.timeslot_combo.pack()
        
        # COLOR CODE FILTER
        color_frame = tk.Frame(filter_content, bg="#f5f5f5")
        color_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(color_frame, text="COLOR CODE", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        self.color_combo = ttk.Combobox(color_frame, textvariable=self.selected_color_code, 
                                      state="readonly", width=20)
        self.color_combo.pack()
        
        # ALGORITHM FILTER
        algorithm_frame = tk.Frame(filter_content, bg="#f5f5f5")
        algorithm_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(algorithm_frame, text="ALGORITHM", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        self.algorithm_combo = ttk.Combobox(algorithm_frame, textvariable=self.selected_algorithm, 
                                         state="readonly", width=20)
        self.algorithm_combo.pack()
        
        # KEY FILTER
        key_frame = tk.Frame(filter_content, bg="#f5f5f5")
        key_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(key_frame, text="KEY", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        self.key_combo = ttk.Combobox(key_frame, textvariable=self.selected_key, 
                                   state="readonly", width=20)
        self.key_combo.pack()
        
        # DATE FILTER
        date_filter_frame = tk.Frame(filter_content, bg="#f5f5f5")
        date_filter_frame.pack(fill=tk.X, pady=(0, 20))
        
        # "From" field with mask
        from_frame = tk.Frame(date_filter_frame, bg="#f5f5f5")
        from_frame.pack(fill=tk.X, pady=2)
        
        from_label = tk.Label(from_frame, text="From:", bg="#f5f5f5", 
                             fg="#000000", font=("Arial", 9), width=5)
        from_label.pack(side=tk.LEFT)
        
        self.date_from_entry = DateMaskEntry(from_frame, textvariable=self.date_from_var,
                                            bg="#f5f5f5")
        self.date_from_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # "To" field with mask
        to_frame = tk.Frame(date_filter_frame, bg="#f5f5f5")
        to_frame.pack(fill=tk.X, pady=2)
        
        to_label = tk.Label(to_frame, text="To:", bg="#f5f5f5",
                           fg="#000000", font=("Arial", 9), width=5)
        to_label.pack(side=tk.LEFT)
        
        self.date_to_entry = DateMaskEntry(to_frame, textvariable=self.date_to_var,
                                          bg="#f5f5f5")
        self.date_to_entry.pack(side=tk.LEFT, padx=(5, 0))
        
        # APPLY and CLEAR buttons after date fields
        button_frame_1 = tk.Frame(date_filter_frame, bg="#f5f5f5")
        button_frame_1.pack(fill=tk.X, pady=(10, 0))
        
        self.apply_button = tk.Button(button_frame_1, text="APPLY", 
                                     command=self.apply_filters,
                                     bg="#4CAF50", fg="white",
                                     font=("Arial", 10, "bold"),
                                     relief=tk.RAISED, bd=2,
                                     cursor="hand2", width=15)
        self.apply_button.pack(fill=tk.X)
        
        # Clear Filter button
        self.clear_button = tk.Button(button_frame_1, text="CLEAR FILTER", 
                                     command=self.clear_filters,
                                     bg="#f44336", fg="white",
                                     font=("Arial", 10, "bold"),
                                     relief=tk.RAISED, bd=2,
                                     cursor="hand2", width=15)
        self.clear_button.pack(fill=tk.X, pady=(5, 0))
        
        # Bottom panel with SESSION GAP and RELOAD DATA
        apply_frame = tk.Frame(filter_frame, bg="#f5f5f5")
        apply_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        
        # SESSION GROUPING GAP
        gap_frame = tk.Frame(apply_frame, bg="#f5f5f5")
        gap_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(gap_frame, text="SESSION GAP", bg="#f5f5f5", 
                 fg="#000000", font=("Arial", 9, "bold")).pack()
        
        gap_input_frame = tk.Frame(gap_frame, bg="#f5f5f5")
        gap_input_frame.pack(fill=tk.X, pady=2)
        tk.Label(gap_input_frame, text="Gap:", bg="#f5f5f5", width=7).pack(side=tk.LEFT)
        tk.Entry(gap_input_frame, textvariable=self.session_gap_var, width=10).pack(side=tk.LEFT)
        tk.Label(gap_input_frame, text="sec", bg="#f5f5f5", font=("Arial", 9)).pack(side=tk.LEFT, padx=(2, 0))
        
        # Optimal value label
        optimal_label = tk.Label(apply_frame, text="Optimal Value 15 sec", 
                                bg="#f5f5f5", fg="#666666",
                                font=("Arial", 8, "italic"))
        optimal_label.pack(pady=(5, 2))
        
        # Reload Data button
        self.reload_button = tk.Button(apply_frame, text="RELOAD DATA", 
                                      command=self.reload_data_with_new_gap,
                                      bg="#FF9800", fg="white",
                                      font=("Arial", 10, "bold"),
                                      relief=tk.RAISED, bd=2,
                                      cursor="hand2", width=15)
        self.reload_button.pack(fill=tk.X, pady=(5, 0))
        
        # Filter status
        self.filter_status = tk.Label(apply_frame, text="No filter applied",
                                     bg="#f5f5f5", fg="#606060",
                                     font=("Arial", 8))
        self.filter_status.pack(pady=(5, 0))
        
        v_separator = tk.Frame(content_frame, width=1, bg="#d0d0d0")
        v_separator.pack(side=tk.LEFT, fill=tk.Y)
        
        # Main content area with scrollbars
        content_area = tk.Frame(content_frame)
        content_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        v_scroll = tk.Scrollbar(content_area, orient=tk.VERTICAL)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        h_scroll = tk.Scrollbar(content_area, orient=tk.HORIZONTAL)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas = tk.Canvas(content_area, bg="#ffffff",
                               yscrollcommand=v_scroll.set,
                               xscrollcommand=h_scroll.set,
                               highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        v_scroll.config(command=self.canvas.yview)
        h_scroll.config(command=self.canvas.xview)
        
        # Bind mouse wheel scrolling
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Button-4>", self._on_mousewheel)  # Linux scroll up
        self.canvas.bind("<Button-5>", self._on_mousewheel)  # Linux scroll down
        
        # Bind scroll events to sync headers
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        
        # Store scrollbar references for synchronization
        self.v_scroll = v_scroll
        self.h_scroll = h_scroll
    
    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling"""
        # Windows and Mac
        if event.delta:
            delta = -1 * (event.delta / 120)
        # Linux
        elif event.num == 4:
            delta = -1
        elif event.num == 5:
            delta = 1
        else:
            delta = 0
        
        # Scroll vertically
        self.canvas.yview_scroll(int(delta), "units")
        
        # Sync left header canvas with main canvas
        try:
            main_scroll = self.canvas.yview()
            if main_scroll:
                self.left_header_canvas.yview_moveto(main_scroll[0])
        except:
            pass
    
    def _on_canvas_configure(self, event):
        """Handle canvas configuration changes"""
        # Update scroll region when canvas size changes
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    
    def _on_vertical_scroll(self, *args):
        """Handle vertical scrolling - sync with left headers"""
        self.canvas.yview(*args)
        # Sync left header canvas with same scroll position
        self.left_header_canvas.yview(*args)
        
        # Additional synchronization to ensure left headers move
        try:
            # Get current scroll position from main canvas
            main_scroll = self.canvas.yview()
            if main_scroll:
                # Apply exact same position to left header canvas
                self.left_header_canvas.yview_moveto(main_scroll[0])
        except:
            pass
        
        # Force update to ensure synchronization
        self.root.update_idletasks()
    
    def _on_horizontal_scroll(self, *args):
        """Handle horizontal scrolling - sync with top header"""
        self.canvas.xview(*args)
        # Sync top header canvas
        self.top_header_canvas.xview(*args)
    
    
    
    def parse_timestamp(self, timestamp_str):
        """Parse timestamp from format YYYY:MM:DD:HH:MM:SS"""
        try:
            return datetime.strptime(timestamp_str, "%Y:%m:%d:%H:%M:%S")
        except:
            return None
    
    def parse_date(self, date_str):
        """Parse date from format DD/MM/YY HH:MM:SS"""
        try:
            return datetime.strptime(date_str, "%d/%m/%y %H:%M:%S")
        except:
            return None
    
    def format_output_timestamp(self, dt):
        """Format datetime for output"""
        return dt.strftime("%Y-%m-%d %H:%M")
    
    def load_all_data(self):
        """Load data from all files"""
        self.file_data.clear()
        self.filtered_file_data.clear()
        
        # Clear unique values
        self.unique_events.clear()
        self.unique_timeslots.clear()
        self.unique_color_codes.clear()
        self.unique_algorithms.clear()
        self.unique_keys.clear()
        
        # Initialize selected files to all files
        self.selected_files = set()
        
        if not os.path.exists(self.input_dir):
            print(f"Directory not found: {self.input_dir}")
            return
        
        # Get all .txt files
        file_paths = []
        for filename in os.listdir(self.input_dir):
            if filename.endswith('.txt'):
                file_path = os.path.join(self.input_dir, filename)
                file_paths.append(file_path)
        
        file_paths.sort()
        
        for file_path in file_paths:
            data = self.load_file_data(file_path)
            if data:
                self.file_data.append(data)
                
                # Collect unique values
                self.unique_events.update(data.get('events', set()))
                self.unique_timeslots.update(data.get('timeslots', set()))
                self.unique_color_codes.update(data.get('color_codes', set()))
                self.unique_algorithms.update(data.get('algorithms', set()))
                self.unique_keys.update(data.get('keys', set()))
        
        # Initialize selected files to all files
        self.selected_files = set(range(len(self.file_data)))
    
    def load_file_data(self, file_path):
        """Load data from single file"""
        sessions = []
        
        # Types of events that should be displayed even without duration
        events_without_duration = {
            'Data Call', 'Data Packet', 'Command', 'Hytera RRS',
            'Motorola LRRP', 'Motorola XCMP', 'Radio Check',
            'Register', 'Response', 'SMS', 'Unit To Unit Call'
        }
        
        # Extract frequency from filename
        frequency = "Unknown"
        try:
            filename = os.path.basename(file_path)
            if '-' in filename:
                parts = filename.replace('.txt', '').split('-')
                if len(parts) == 3:
                    freq_mhz = parts[0]
                    freq_khz = parts[1] + parts[2]
                    frequency = f"{freq_mhz}.{freq_khz} MHz"
        except:
            pass
        
        # Unique values for filters
        events = set()
        timeslots = set()
        color_codes = set()
        algorithms = set()
        keys = set()
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    event_type = row.get('EVENT', '')
                    has_duration = row.get('DURATION_MS') and row['DURATION_MS'] != ''
                    
                    # Collect unique values
                    if event_type:
                        events.add(event_type)
                    if row.get('TIMESLOT'):
                        timeslots.add(row.get('TIMESLOT'))
                    if row.get('COLOR_CODE'):
                        color_codes.add(row.get('COLOR_CODE'))
                    if row.get('ALGORITHM'):
                        algorithms.add(row.get('ALGORITHM'))
                    if row.get('KEY'):
                        keys.add(row.get('KEY'))
                    
                    # Process records with duration
                    if has_duration:
                        try:
                            duration_ms = int(row['DURATION_MS'])
                        except ValueError:
                            continue
                        
                        if duration_ms < 500:  # Skip sessions shorter than 0.5 seconds
                            continue
                        
                        # Parse data
                        timestamp = self.parse_timestamp(row['TIMESTAMP'])
                        duration_sec = duration_ms / 1000.0
                        
                        # Extract necessary fields
                        session_data = {
                            'timestamp': timestamp,
                            'end_time': timestamp + timedelta(milliseconds=duration_ms) if timestamp else None,
                            'duration_sec': duration_sec,
                            'duration_ms': duration_ms,
                            'timeslot': row.get('TIMESLOT', ''),
                            'color_code': row.get('COLOR_CODE', ''),
                            'from': row.get('FROM', ''),
                            'to': row.get('TO', ''),
                            'event': event_type,
                            'algorithm': row.get('ALGORITHM', ''),
                            'key': row.get('KEY', ''),
                            'has_duration': True
                        }
                        
                        sessions.append(session_data)
                    
                    # Process records without duration for certain event types
                    elif event_type in events_without_duration:
                        timestamp = self.parse_timestamp(row['TIMESTAMP'])
                        
                        session_data = {
                            'timestamp': timestamp,
                            'end_time': timestamp,  # For events without duration
                            'duration_sec': 0,
                            'duration_ms': 0,
                            'timeslot': row.get('TIMESLOT', ''),
                            'color_code': row.get('COLOR_CODE', ''),
                            'from': row.get('FROM', ''),
                            'to': row.get('TO', ''),
                            'event': event_type,
                            'algorithm': row.get('ALGORITHM', ''),
                            'key': row.get('KEY', ''),
                            'has_duration': False
                        }
                        
                        sessions.append(session_data)
        
        except FileNotFoundError:
            print(f"File not found: {file_path}")
            return None
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            return None
        
        # Sort sessions by time
        sessions.sort(key=lambda x: x['timestamp'] if x['timestamp'] else datetime.min)
        
        # Group sessions (combine if gap between them is less than specified seconds)
        grouped_sessions = []
        current_group = []
        try:
            gap_between_recording_sec = int(self.session_gap_var.get())
        except:
            gap_between_recording_sec = 15  # Default value
        
        for session in sessions:
            if not current_group:
                current_group.append(session)
            else:
                # Check time between end of last session and start of current
                last_session = current_group[-1]
                if last_session['end_time'] and session['timestamp']:
                    time_gap = (session['timestamp'] - last_session['end_time']).total_seconds()
                    
                    if time_gap <= gap_between_recording_sec:
                        current_group.append(session)
                    else:
                        # Save current group and start new
                        grouped_sessions.append(current_group)
                        current_group = [session]
                else:
                    current_group.append(session)
        
        # Add last group
        if current_group:
            grouped_sessions.append(current_group)
        
        return {
            'file_path': file_path,
            'filename': os.path.basename(file_path),
            'frequency': frequency,
            'sessions': grouped_sessions,
            'events': events,
            'timeslots': timeslots,
            'color_codes': color_codes,
            'algorithms': algorithms,
            'keys': keys
        }
    
    def update_comboboxes(self):
        """Update comboboxes with unique values"""
        # EVENT
        event_values = ['All'] + sorted(self.unique_events)
        self.event_combo['values'] = event_values
        if self.selected_event.get() not in event_values:
            self.selected_event.set('All')
        
        # TIMESLOT
        timeslot_values = ['All'] + sorted(self.unique_timeslots)
        self.timeslot_combo['values'] = timeslot_values
        if self.selected_timeslot.get() not in timeslot_values:
            self.selected_timeslot.set('All')
        
        # COLOR CODE
        color_values = ['All'] + sorted(self.unique_color_codes)
        self.color_combo['values'] = color_values
        if self.selected_color_code.get() not in color_values:
            self.selected_color_code.set('All')
        
        # ALGORITHM
        algorithm_values = ['All'] + sorted(self.unique_algorithms)
        self.algorithm_combo['values'] = algorithm_values
        if self.selected_algorithm.get() not in algorithm_values:
            self.selected_algorithm.set('All')
        
        # KEY
        key_values = ['All'] + sorted(self.unique_keys)
        self.key_combo['values'] = key_values
        if self.selected_key.get() not in key_values:
            self.selected_key.set('All')
    
    def apply_filters(self):
        """Apply filters to data"""
        date_from_str = self.date_from_var.get()
        date_to_str = self.date_to_var.get()
        
        date_from = self.parse_date(date_from_str)
        date_to = self.parse_date(date_to_str)
        
        if not date_from or not date_to:
            self.filter_status.config(text="Invalid date format", fg="red")
            return
        
        # Get filter values
        duration_from = self.duration_from_var.get()
        duration_to = self.duration_to_var.get()
        event_filter = self.selected_event.get()
        timeslot_filter = self.selected_timeslot.get()
        color_code_filter = self.selected_color_code.get()
        algorithm_filter = self.selected_algorithm.get()
        key_filter = self.selected_key.get()
        
        # Convert duration from seconds to milliseconds
        try:
            dur_from_ms = float(duration_from) * 1000 if duration_from else None
            dur_to_ms = float(duration_to) * 1000 if duration_to else None
        except:
            dur_from_ms = dur_to_ms = None
        
        # Filter data
        self.filtered_file_data = []
        total_sessions = 0
        total_filtered = 0
        
        for data in self.file_data:
            filtered_sessions = []
            
            for group in data['sessions']:
                filtered_group = []
                
                for session in group:
                    total_sessions += 1
                    passes_filter = True
                    
                    # Filter by date
                    if session['timestamp']:
                        if not (date_from <= session['timestamp'] <= date_to):
                            passes_filter = False
                    else:
                        passes_filter = False
                    
                    # Filter by duration
                    if dur_from_ms is not None or dur_to_ms is not None:
                        duration_ms = session.get('duration_ms', 0)
                        if dur_from_ms is not None and duration_ms < dur_from_ms:
                            passes_filter = False
                        if dur_to_ms is not None and duration_ms > dur_to_ms:
                            passes_filter = False
                    
                    # Filter by EVENT
                    if event_filter != 'All' and session['event'] != event_filter:
                        passes_filter = False
                    
                    # Filter by TIMESLOT
                    if timeslot_filter != 'All' and session['timeslot'] != timeslot_filter:
                        passes_filter = False
                    
                    # Filter by COLOR_CODE
                    if color_code_filter != 'All' and session['color_code'] != color_code_filter:
                        passes_filter = False
                    
                    # Filter by ALGORITHM
                    if algorithm_filter != 'All' and session['algorithm'] != algorithm_filter:
                        passes_filter = False
                    
                    # Filter by KEY
                    if key_filter != 'All' and session['key'] != key_filter:
                        passes_filter = False
                    
                    if passes_filter:
                        filtered_group.append(session)
                        total_filtered += 1
                
                if filtered_group:
                    filtered_sessions.append(filtered_group)
            
            filtered_data = {
                'file_path': data['file_path'],
                'filename': data['filename'],
                'frequency': data['frequency'],
                'sessions': filtered_sessions,
                'events': data['events'],
                'timeslots': data['timeslots'],
                'color_codes': data['color_codes'],
                'algorithms': data['algorithms'],
                'keys': data['keys']
            }
            self.filtered_file_data.append(filtered_data)
        
        # Update status
        if total_sessions > 0:
            percent = int((total_filtered / total_sessions) * 100) if total_sessions > 0 else 0
            self.filter_status.config(
                text=f"Filtered: {total_filtered}/{total_sessions} ({percent}%)",
                fg="green" if total_filtered > 0 else "orange"
            )
        else:
            self.filter_status.config(
                text="No data found",
                fg="red"
            )
        
        # Redraw with filtered data
        self.display_data(use_filtered=True)
    
    def clear_filters(self):
        """Clear all filters"""
        self.filtered_file_data = []
        
        # Reset filter values
        self.duration_from_var.set("")
        self.duration_to_var.set("")
        self.selected_event.set("All")
        self.selected_timeslot.set("All")
        self.selected_color_code.set("All")
        self.selected_algorithm.set("All")
        self.selected_key.set("All")
        
        # Update status
        self.filter_status.config(text="No filter applied", fg="#606060")
        
        # Redraw without filters
        self.display_data(use_filtered=False)
    
    def display_data(self, use_filtered=None):
        """Display data in Excel-like table format"""
        print(f"=== display_data called with use_filtered={use_filtered} ===")
        print(f"Current selected_files: {self.selected_files}")
        print(f"selected_files type: {type(self.selected_files)}")
        print(f"selected_files length: {len(self.selected_files) if self.selected_files else 'None'}")
        
        self.canvas.delete("all")
        
        # If use_filtered not specified, determine automatically
        if use_filtered is None:
            use_filtered = len(self.filtered_file_data) > 0
        
        # Choose data to display
        data_to_display = self.filtered_file_data if use_filtered and self.filtered_file_data else self.file_data
        print(f"Initial data_to_display count: {len(data_to_display)}")
        
        # Filter by selected files if any are selected
        print(f"Before filtering: selected_files={self.selected_files}, data_to_display count={len(data_to_display)}")
        
        if self.selected_files:
            data_to_display = [data for idx, data in enumerate(data_to_display) if idx in self.selected_files]
            print(f"Filtering by selected files: {self.selected_files}, showing {len(data_to_display)} files")  # Debug info
        else:
            # If no files are selected, show empty list (no files to display)
            data_to_display = []
            print(f"No files selected, showing empty list")  # Debug info
            print(f"self.selected_files is empty: {len(self.selected_files) == 0}")
            print(f"self.selected_files content: {self.selected_files}")
        
        print(f"Final data_to_display count: {len(data_to_display)}")
        
        if not data_to_display:
            self.canvas.create_text(400, 200, text="No data to display", 
                                   font=("Arial", 14), fill="#000000")
            return
        
        # Separate files with data and empty files
        files_with_data = []
        empty_files = []
        
        for data in data_to_display:
            if data['sessions']:
                files_with_data.append(data)
            else:
                empty_files.append(data)
        
        # If no files to display, show empty message
        if not files_with_data and not empty_files:
            print(f"No files to display - files_with_data: {len(files_with_data)}, empty_files: {len(empty_files)}")
            self.canvas.create_text(400, 200, text="No files selected", 
                                   font=("Arial", 14), fill="#000000")
            return
        
        print(f"Displaying {len(files_with_data)} files with data and {len(empty_files)} empty files")
        
        # Create time-based data structure
        time_data = self.create_time_based_data(files_with_data)
        
        # Calculate layout
        cell_width = 200  # Increased width for better content display
        base_cell_height = 30
        date_column_width = 100
        hour_column_width = 30
        start_x = 20
        start_y = 20
        
        # Draw everything on main canvas (no fixed headers)
        current_y = start_y  # Start from top with margin
        
        # Draw frequency headers on main canvas
        header_y = current_y
        current_x = start_x + date_column_width + hour_column_width
        
        for data in files_with_data:
            # Draw header cell
            self.canvas.create_rectangle(
                current_x, header_y,
                current_x + cell_width, header_y + base_cell_height,
                outline="#000000", width=1, fill="#e0e0e0"
            )
            
            # File name (truncated)
            filename = data['filename'][:12] + "..." if len(data['filename']) > 12 else data['filename']
            self.canvas.create_text(
                current_x + cell_width // 2, header_y + 10,
                text=filename,
                font=("Arial", 8, "bold"),
                fill="#000000"
            )
            
            # Frequency
            self.canvas.create_text(
                current_x + cell_width // 2, header_y + 22,
                text=data['frequency'],
                font=("Arial", 7),
                fill="#666666"
            )
            
            current_x += cell_width
        
        # Draw empty file headers
        for data in empty_files:
            self.canvas.create_rectangle(
                current_x, header_y,
                current_x + cell_width, header_y + base_cell_height,
                outline="#cccccc", width=1, fill="#f5f5f5"
            )
            
            filename = data['filename'][:12] + "..." if len(data['filename']) > 12 else data['filename']
            self.canvas.create_text(
                current_x + cell_width // 2, header_y + 10,
                text=filename,
                font=("Arial", 8),
                fill="#999999"
            )
            
            self.canvas.create_text(
                current_x + cell_width // 2, header_y + 22,
                text=data['frequency'],
                font=("Arial", 7),
                fill="#cccccc"
            )
            
            current_x += cell_width
        
        # Move to next row after headers
        current_y += base_cell_height + 10
        
        all_dates_data = []
        
        for date_str, hours_data in time_data.items():
            # Find the actual range of hours with data
            hours_with_data = set()
            for hour in range(24):
                for data in files_with_data:
                    hour_sessions = hours_data.get(hour, {}).get(data['filename'], [])
                    if hour_sessions:
                        hours_with_data.add(hour)
            
            if not hours_with_data:
                continue  # Skip dates with no data
            
            # Calculate total height needed for this date
            max_hour_height = 0
            
            # First pass: calculate heights for each hour with data
            hour_heights = {}
            for hour in sorted(hours_with_data):
                hour_height = base_cell_height
                
                # Check if any cell in this hour has sessions
                for data in files_with_data:
                    hour_sessions = hours_data.get(hour, {}).get(data['filename'], [])
                    if hour_sessions:
                        cell_key = (date_str, hour, data['filename'])
                        # Calculate height needed for content (always show groups)
                        content_height = self.calculate_cell_content_height(hour_sessions, cell_key)
                        hour_height = max(hour_height, content_height)
                
                hour_heights[hour] = hour_height
                max_hour_height = max(max_hour_height, hour_height)
            
            # Calculate total height for this date (sum of actual hour heights)
            date_height = sum(hour_heights[hour] for hour in hours_with_data)
            
            # Store data for left headers
            all_dates_data.append({
                'date': date_str,
                'height': date_height,
                'hours': hours_with_data,
                'hour_heights': hour_heights
            })
            
            # Draw date and hour headers on main canvas
            hour_y = current_y
            for hour in sorted(hours_with_data):
                hour_height = hour_heights[hour]
                
                # Draw date header (only for first hour of each date)
                if hour == min(hours_with_data):
                    self.canvas.create_rectangle(
                        start_x, current_y,
                        start_x + date_column_width, current_y + date_height,
                        outline="#000000", width=1, fill="#f0f0f0"
                    )
                    
                    self.canvas.create_text(
                        start_x + date_column_width // 2, current_y + date_height // 2,
                        text=date_str,
                        font=("Arial", 10, "bold"),
                        fill="#000000"
                    )
                
                # Draw hour header
                self.canvas.create_rectangle(
                    start_x + date_column_width, hour_y,
                    start_x + date_column_width + hour_column_width, hour_y + hour_height,
                    outline="#000000", width=1, fill="#f8f8f8"
                )
                
                self.canvas.create_text(
                    start_x + date_column_width + hour_column_width // 2, hour_y + hour_height // 2,
                    text=f"{hour:02d}",
                    font=("Arial", 9, "bold"),
                    fill="#000000"
                )
                
                # Draw data cells for this hour
                cell_x = start_x + date_column_width + hour_column_width  # Start after date and hour columns
                
                for data in files_with_data:
                    # Get sessions for this hour
                    hour_sessions = hours_data.get(hour, {}).get(data['filename'], [])
                    cell_key = (date_str, hour, data['filename'])
                    
                    # Determine cell color based on session count
                    if hour_sessions:
                        session_count = sum(len(group) for group in hour_sessions)
                        if session_count > 10:
                            fill_color = "#ffcccc"  # Light red for many sessions
                        elif session_count > 5:
                            fill_color = "#ffffcc"  # Light yellow for medium sessions
                        else:
                            fill_color = "#ccffcc"  # Light green for few sessions
                    else:
                        fill_color = "#ffffff"  # White for no sessions
                    
                    # Draw cell
                    cell_rect = self.canvas.create_rectangle(
                        cell_x, hour_y,
                        cell_x + cell_width, hour_y + hour_height,
                        outline="#cccccc", width=1, fill=fill_color,
                        tags=f"cell_{date_str}_{hour}_{data['filename']}"
                    )
                    
                    # Display content in cell
                    if hour_sessions:
                        self.draw_cell_content(cell_x, hour_y, cell_width, hour_height, 
                                             hour_sessions, cell_key, True)
                    
                    cell_x += cell_width
                
                # Draw empty cells for empty files
                for data in empty_files:
                    self.canvas.create_rectangle(
                        cell_x, hour_y,
                        cell_x + cell_width, hour_y + hour_height,
                        outline="#cccccc", width=1, fill="#f9f9f9"
                    )
                    cell_x += cell_width
                
                hour_y += hour_height
            
            current_y += date_height + 10  # Add spacing between dates
        
        # Update scroll region
        total_width = len(data_to_display) * cell_width + start_x + date_column_width + hour_column_width
        total_height = current_y + 20
        self.canvas.configure(scrollregion=(0, 0, total_width, total_height))
    
    def create_time_based_data(self, files_with_data):
        """Create time-based data structure grouped by date and hour"""
        time_data = {}
        
        for data in files_with_data:
            for group in data['sessions']:
                if not group or not group[0]['timestamp']:
                    continue
                
                timestamp = group[0]['timestamp']
                date_str = timestamp.strftime("%Y-%m-%d")
                hour = timestamp.hour
                
                if date_str not in time_data:
                    time_data[date_str] = {}
                
                if hour not in time_data[date_str]:
                    time_data[date_str][hour] = {}
                
                if data['filename'] not in time_data[date_str][hour]:
                    time_data[date_str][hour][data['filename']] = []
                
                time_data[date_str][hour][data['filename']].append(group)
        
        return time_data
    
    def calculate_cell_content_height(self, hour_sessions, cell_key):
        """Calculate height needed for cell content"""
        if not hour_sessions:
            return 30
        
        line_height = 12
        padding = 10
        base_height = 30
        
        # Always show group headers, calculate height based on expanded groups
        total_height = padding
        
        for group_index, group in enumerate(hour_sessions):
            if not group:
                continue
            
            # Group header (always shown)
            total_height += line_height
            
            # Check if this group is expanded
            group_key = (*cell_key, group_index)
            if group_key in self.expanded_groups:
                # Show sessions in group
                for session in group:
                    total_height += line_height
        
        return max(base_height, total_height + padding)
    
    def draw_cell_content(self, x, y, width, height, hour_sessions, cell_key, is_expanded):
        """Draw content inside cell"""
        if not hour_sessions:
            return
        
        line_height = 12
        padding = 5
        current_y = y + padding
        
        # Always show groups (collapsed by default)
        for group_index, group in enumerate(hour_sessions):
            if not group:
                continue
            
            # Group header
            timestamp_str = self.format_output_timestamp(group[0]['timestamp']) if group[0]['timestamp'] else "Unknown"
            group_key = (*cell_key, group_index)
            is_group_expanded = group_key in self.expanded_groups
            
            # Extract variables from cell_key
            date_str, hour, filename = cell_key
            
            # Add expand/collapse indicator
            indicator = "▼" if is_group_expanded else "▶"
            group_header = f"{indicator} {timestamp_str} ({len(group)})"
             
            # Truncate if too long
            if len(group_header) > 25:
                group_header = group_header[:22] + "..."
            
            # Create a safe tag name (replace special characters)
            safe_filename = filename.replace("/", "_").replace("\\", "_").replace(":", "_").replace(" ", "_")
            safe_tag = f"group_header_{date_str}_{hour}_{safe_filename}_{group_index}"
            
            # Draw clickable group header
            header_text = self.canvas.create_text(
                x + padding, current_y,
                text=group_header,
                font=("Arial", 8, "bold"),
                fill="#0000ff",
                anchor="w",
                tags=safe_tag
            )
            
            # Add click event to group header
            def on_group_click(event, key=group_key):
                print(f"Group clicked: {key}")  # Debug info
                self.toggle_group_collapse_in_cell(key)
            
            self.canvas.tag_bind(safe_tag, "<Button-1>", on_group_click)
            
            current_y += line_height
            
            # Show sessions if group is expanded
            if is_group_expanded:
                for session in group:
                    if current_y + line_height > y + height - padding:
                        break  # Don't overflow cell
                    
                    # Format session line
                    line_parts = []
                    
                    if session['timeslot']:
                        line_parts.append(session['timeslot'])
                    
                    if session['color_code']:
                        line_parts.append(f"CC{session['color_code']}")
                    
                    if session['from'] and session['to']:
                        line_parts.append(f"{session['from']}→{session['to']}")
                    
                    if session['has_duration']:
                        line_parts.append(f"({session['duration_sec']:.1f}s)")
                    
                    if session['event']:
                        line_parts.append(session['event'])
                    
                    session_line = " ".join(line_parts)
                    
                    # Truncate if too long
                    if len(session_line) > 30:
                        session_line = session_line[:27] + "..."
                    
                    self.canvas.create_text(
                        x + padding + 10, current_y,
                        text=session_line,
                        font=("Courier", 7),
                        fill="#000000",
                        anchor="w"
                    )
                    
                    current_y += line_height
    
    
    def toggle_group_collapse_in_cell(self, group_key):
        """Toggle collapse/expand state of a group within a cell"""
        print(f"Toggling group: {group_key}")  # Debug info
        
        if group_key in self.expanded_groups:
            self.expanded_groups.remove(group_key)
            print(f"Group collapsed: {group_key}")
        else:
            self.expanded_groups.add(group_key)
            print(f"Group expanded: {group_key}")
        
        # Redraw the table
        self.display_data()
    
    def draw_fixed_frequency_headers(self, files_with_data, empty_files, cell_width, base_cell_height, start_x, date_column_width, hour_column_width):
        """Draw fixed frequency headers at the top"""
        # Clear previous headers
        self.top_header_canvas.delete("all")
        
        # Draw frequency headers
        header_y = 10
        current_x = start_x + date_column_width + hour_column_width
        
        for data in files_with_data:
            # Draw header cell
            self.top_header_canvas.create_rectangle(
                current_x, header_y,
                current_x + cell_width, header_y + base_cell_height,
                outline="#000000", width=1, fill="#e0e0e0"
            )
            
            # File name (truncated)
            filename = data['filename'][:12] + "..." if len(data['filename']) > 12 else data['filename']
            self.top_header_canvas.create_text(
                current_x + cell_width // 2, header_y + 10,
                text=filename,
                font=("Arial", 8, "bold"),
                fill="#000000"
            )
            
            # Frequency
            self.top_header_canvas.create_text(
                current_x + cell_width // 2, header_y + 22,
                text=data['frequency'],
                font=("Arial", 7),
                fill="#666666"
            )
            
            current_x += cell_width
        
        # Draw empty file headers
        for data in empty_files:
            self.top_header_canvas.create_rectangle(
                current_x, header_y,
                current_x + cell_width, header_y + base_cell_height,
                outline="#cccccc", width=1, fill="#f5f5f5"
            )
            
            filename = data['filename'][:12] + "..." if len(data['filename']) > 12 else data['filename']
            self.top_header_canvas.create_text(
                current_x + cell_width // 2, header_y + 10,
                text=filename,
                font=("Arial", 8),
                fill="#999999"
            )
            
            self.top_header_canvas.create_text(
                current_x + cell_width // 2, header_y + 22,
                text=data['frequency'],
                font=("Arial", 7),
                fill="#cccccc"
            )
            
            current_x += cell_width
        
        # Set scroll region for top header canvas
        self.top_header_canvas.configure(scrollregion=(0, 0, current_x, base_cell_height + 20))
    
    def draw_fixed_left_headers(self, all_dates_data):
        """Draw fixed date and hour headers on the left"""
        # Clear previous left headers
        self.left_header_canvas.delete("all")
        
        current_y = 20
        date_column_width = 100
        hour_column_width = 40
        
        for date_data in all_dates_data:
            date_str = date_data['date']
            date_height = date_data['height']
            hours_with_data = date_data['hours']
            hour_heights = date_data['hour_heights']
            
            # Date header
            self.left_header_canvas.create_rectangle(
                10, current_y,
                10 + date_column_width, current_y + date_height,
                outline="#000000", width=1, fill="#f0f0f0"
            )
            
            self.left_header_canvas.create_text(
                10 + date_column_width // 2, current_y + date_height // 2,
                text=date_str,
                font=("Arial", 10, "bold"),
                fill="#000000",
                angle=90
            )
            
            # Draw hour labels
            hour_y = current_y
            for hour in sorted(hours_with_data):
                hour_height = hour_heights[hour]
                
                # Hour label
                self.left_header_canvas.create_rectangle(
                    10 + date_column_width, hour_y,
                    10 + date_column_width + hour_column_width, hour_y + hour_height,
                    outline="#000000", width=1, fill="#f8f8f8"
                )
                
                self.left_header_canvas.create_text(
                    10 + date_column_width + hour_column_width // 2, hour_y + hour_height // 2,
                    text=f"{hour:02d}",
                    font=("Arial", 9, "bold"),
                    fill="#000000"
                )
                
                hour_y += hour_height
            
            current_y += date_height + 10
        
        # Set scroll region for left header canvas to match main canvas height
        self.left_header_canvas.configure(scrollregion=(0, 0, date_column_width + hour_column_width + 20, current_y + 20))
        
        # Force synchronization with main canvas
        self.sync_left_headers_with_main()
    
    def sync_left_headers_with_main(self):
        """Force synchronization of left headers with main canvas"""
        try:
            # Get main canvas scroll position
            main_scroll_info = self.canvas.yview()
            if main_scroll_info:
                # Apply same scroll position to left header canvas
                self.left_header_canvas.yview_moveto(main_scroll_info[0])
        except Exception as e:
            print(f"Sync error: {e}")
    
    
    def set_date_range_from_data(self):
        """Set date range based on loaded data"""
        min_date = None
        max_date = None
        
        # Search for minimum and maximum dates in all files
        for data in self.file_data:
            for group in data['sessions']:
                for session in group:
                    if session['timestamp']:
                        if min_date is None or session['timestamp'] < min_date:
                            min_date = session['timestamp']
                        if max_date is None or session['timestamp'] > max_date:
                            max_date = session['timestamp']
        
        # If dates found, set them
        if min_date and max_date:
            # Format dates to required format DD/MM/YY HH:MM:SS
            from_str = min_date.strftime("%d/%m/%y %H:%M:%S")
            to_str = max_date.strftime("%d/%m/%y %H:%M:%S")
            
            self.date_from_var.set(from_str)
            self.date_to_var.set(to_str)
            
            print(f"Date range from data: {from_str} to {to_str}")
        else:
            print("No dates found in data, using default range")
    
    def refresh_all(self):
        """Refresh all data"""
        self.load_all_data()
        self.filtered_file_data = []
        
        # Update date range from new data
        self.set_date_range_from_data()
        
        # Update comboboxes
        self.update_comboboxes()
        
        self.filter_status.config(text="No filter applied", fg="#606060")
        self.display_data()
    
    def reload_data_with_new_gap(self):
        """Reload data with new session gap setting"""
        try:
            gap_value = int(self.session_gap_var.get())
            if gap_value < 1 or gap_value > 300:  # Reasonable limits
                messagebox.showwarning("Invalid Gap", "Session gap must be between 1 and 300 seconds")
                return
        except ValueError:
            messagebox.showwarning("Invalid Gap", "Please enter a valid number for session gap")
            return
        
        # Reload all data with new gap
        self.load_all_data()
        self.filtered_file_data = []
        
        # Update date range from new data
        self.set_date_range_from_data()
        
        # Update comboboxes
        self.update_comboboxes()
        
        self.filter_status.config(text=f"Data reloaded with {gap_value}s gap", fg="#4CAF50")
        self.display_data()
    
    def open_file_selector(self):
        """Open file selection window"""
        # Store reference to selector window
        self.file_selector_window = tk.Toplevel(self.root)
        self.file_selector_window.title("Select Files")
        self.file_selector_window.geometry("600x700")  # Doubled height from 350 to 700, increased width
        self.file_selector_window.resizable(True, True)  # Allow resizing
        
        # Main container
        main_container = tk.Frame(self.file_selector_window, bg="#f5f5f5")
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Info label showing number of files
        info_label = tk.Label(main_container, 
                            text=f"Total files: {len(self.file_data)}", 
                            bg="#f5f5f5", font=("Arial", 10, "bold"))
        info_label.pack(pady=(0, 5))
        
        # List frame
        files_frame = tk.Frame(main_container, bg="#f5f5f5")
        files_frame.pack(fill=tk.BOTH, expand=True)
        
        # Container for scrolling with better configuration
        canvas = tk.Canvas(files_frame, bg="white")
        scrollbar = tk.Scrollbar(files_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="white")
        
        # Configure canvas scrolling
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create window in canvas
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Update scroll region when frame changes
        def configure_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        scrollable_frame.bind("<Configure>", configure_scroll_region)
        
        # Update canvas width when canvas size changes
        def configure_canvas_width(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        canvas.bind("<Configure>", configure_canvas_width)
        
        # Create header frame with sortable columns
        header_frame = tk.Frame(scrollable_frame, bg="#f0f0f0", relief=tk.RAISED, bd=1)
        header_frame.pack(fill=tk.X, padx=5, pady=(0, 5))
        
        # Sort variables (use existing ones from __init__)
        # Don't create new variables here
        
        # Header buttons with sort indicators
        name_text = "📁 File Name"
        if self.sort_by_name.get() == "asc":
            name_text += " ↑"
        elif self.sort_by_name.get() == "desc":
            name_text += " ↓"
        
        name_header = tk.Button(header_frame, text=name_text, 
                               command=lambda: self.toggle_sort("name"),
                               bg="#e0e0e0", fg="#000000",
                               font=("Arial", 9, "bold"),
                               relief=tk.RAISED, bd=1, width=25)
        name_header.pack(side=tk.LEFT, padx=2, pady=2)
        
        sessions_text = "📊 Sessions"
        if self.sort_by_sessions.get() == "asc":
            sessions_text += " ↑"
        elif self.sort_by_sessions.get() == "desc":
            sessions_text += " ↓"
        
        sessions_header = tk.Button(header_frame, text=sessions_text, 
                                   command=lambda: self.toggle_sort("sessions"),
                                   bg="#e0e0e0", fg="#000000",
                                   font=("Arial", 9, "bold"),
                                   relief=tk.RAISED, bd=1, width=15)
        sessions_header.pack(side=tk.LEFT, padx=2, pady=2)
        
        # Create checkboxes for each file
        checkboxes = []
        checkbox_vars = []
        checkbox_to_file_index = []  # Map checkbox index to file index
        
        # Prepare data for sorting
        file_data_with_counts = []
        for idx, data in enumerate(self.file_data):
            session_count = sum(len(group) for group in data['sessions'])
            file_data_with_counts.append({
                'index': idx,
                'data': data,
                'session_count': session_count
            })
        
        # Sort data based on current sort settings
        sorted_data = self.sort_file_data(file_data_with_counts)
        
        for item in sorted_data:
            var = tk.BooleanVar(value=(item['index'] in self.selected_files))
            checkbox_vars.append(var)
            checkbox_to_file_index.append(item['index'])  # Store mapping
            
            cb_frame = tk.Frame(scrollable_frame, bg="white")
            cb_frame.pack(fill=tk.X, padx=5, pady=2)
            
            # Add file number and session count for easier navigation
            cb_text = f"{item['index'] + 1}. {item['data']['filename']} ({item['session_count']} sessions)"
            
            cb = tk.Checkbutton(cb_frame, text=cb_text,
                            variable=var, bg="white",
                            font=("Arial", 9), anchor="w", 
                            wraplength=550)  # Wrap long filenames
            cb.pack(fill=tk.X)
            checkboxes.append(cb)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        # Bind mouse wheel to canvas
        canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Control buttons
        button_frame = tk.Frame(main_container, bg="#f5f5f5")
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Select all / Deselect all buttons
        control_frame = tk.Frame(button_frame, bg="#f5f5f5")
        control_frame.pack(fill=tk.X, pady=(0, 10))
        
        def deselect_all():
            for var in checkbox_vars:
                var.set(False)
            update_selection_count()
        
        def select_all():
            for var in checkbox_vars:
                var.set(True)
            update_selection_count()
        
        def invert_selection():
            for var in checkbox_vars:
                var.set(not var.get())
            update_selection_count()
        
        # Selection counter
        selection_label = tk.Label(control_frame, text="", bg="#f5f5f5", 
                                font=("Arial", 9))
        selection_label.pack(pady=(0, 5))
        
        def update_selection_count():
            count = sum(1 for var in checkbox_vars if var.get())
            selection_label.config(text=f"Selected: {count} / {len(self.file_data)}")
        
        # Initial count
        update_selection_count()
        
        # Add trace to update count when checkboxes change
        for var in checkbox_vars:
            var.trace('w', lambda *args: update_selection_count())
        
        # Button row 1
        button_row1 = tk.Frame(control_frame, bg="#f5f5f5")
        button_row1.pack(fill=tk.X, pady=2)
        
        deselect_btn = tk.Button(button_row1, text="Deselect All",
                            command=deselect_all,
                            bg="#e0e0e0", font=("Arial", 9),
                            width=12, height=2)
        deselect_btn.pack(side=tk.LEFT, padx=5)
        
        select_all_btn = tk.Button(button_row1, text="Select All",
                                command=select_all,
                                bg="#e0e0e0", font=("Arial", 9),
                                width=12, height=2)
        select_all_btn.pack(side=tk.LEFT, padx=5)
        
        invert_btn = tk.Button(button_row1, text="Invert Selection",
                            command=invert_selection,
                            bg="#e0e0e0", font=("Arial", 9),
                            width=12, height=2)
        invert_btn.pack(side=tk.LEFT, padx=5)
        
        # Apply and Cancel buttons
        button_row2 = tk.Frame(button_frame, bg="#f5f5f5")
        button_row2.pack(fill=tk.X, pady=(10, 0))
        
        def apply_file_selection():
            print(f"Before clearing: selected_files={self.selected_files}")
            self.selected_files.clear()
            print(f"After clearing: selected_files={self.selected_files}")
            
            for checkbox_idx, var in enumerate(checkbox_vars):
                if var.get():
                    file_idx = checkbox_to_file_index[checkbox_idx]
                    self.selected_files.add(file_idx)
                    print(f"Added file {file_idx} (checkbox {checkbox_idx}) to selection")
            
            print(f"Final selected files: {self.selected_files}")  # Debug info
            print(f"Final selected files type: {type(self.selected_files)}")
            print(f"Final selected files length: {len(self.selected_files)}")
            
            # Unbind mousewheel before closing
            canvas.unbind_all("<MouseWheel>")
            
            self.file_selector_window.destroy()
            # Update filter status
            if self.selected_files:
                self.filter_status.config(text=f"File filter: {len(self.selected_files)} files selected", fg="#4CAF50")
            else:
                self.filter_status.config(text="No files selected - showing empty", fg="#f44336")
            
            print(f"About to call display_data with selected_files={self.selected_files}")
            print(f"About to call display_data with selected_files type: {type(self.selected_files)}")
            # Force refresh of display with current file selection
            self.display_data(use_filtered=False)
        
        def cancel_selection():
            # Unbind mousewheel before closing
            canvas.unbind_all("<MouseWheel>")
            self.file_selector_window.destroy()
        
        apply_btn = tk.Button(button_row2, text="APPLY",
                        command=apply_file_selection,
                        bg="#4CAF50", fg="white",
                        font=("Arial", 10, "bold"),
                        width=15, height=2)
        apply_btn.pack(side=tk.LEFT, padx=10)
        
        cancel_btn = tk.Button(button_row2, text="CANCEL",
                            command=cancel_selection,
                            bg="#f44336", fg="white",
                            font=("Arial", 10, "bold"),
                            width=15, height=2)
        cancel_btn.pack(side=tk.RIGHT, padx=10)
        
        # Make window modal
        self.file_selector_window.transient(self.root)
        self.file_selector_window.grab_set()
        
        # Center window
        self.file_selector_window.update_idletasks()
        x = (self.file_selector_window.winfo_screenwidth() // 2) - 300
        y = (self.file_selector_window.winfo_screenheight() // 2) - 350
        self.file_selector_window.geometry(f"+{x}+{y}")
        
        # Focus on first checkbox for keyboard navigation
        if checkboxes:
            checkboxes[0].focus_set()
    
    def toggle_sort(self, sort_type):
        """Toggle sorting for file selector"""
        if sort_type == "name":
            if self.sort_by_name.get() == "asc":
                self.sort_by_name.set("desc")
            else:
                self.sort_by_name.set("asc")
            self.sort_by_sessions.set("none")
        elif sort_type == "sessions":
            if self.sort_by_sessions.get() == "none":
                self.sort_by_sessions.set("desc")
            elif self.sort_by_sessions.get() == "desc":
                self.sort_by_sessions.set("asc")
            else:
                self.sort_by_sessions.set("none")
            self.sort_by_name.set("asc")
        
        # Update the current file selector window
        self.update_file_selector_sorting()
    
    def update_file_selector_sorting(self):
        """Update the file selector window with new sorting"""
        if hasattr(self, 'file_selector_window') and self.file_selector_window.winfo_exists():
            # Close current window and reopen with new sorting
            self.file_selector_window.destroy()
            self.open_file_selector()
    
    def sort_file_data(self, file_data_with_counts):
        """Sort file data based on current sort settings"""
        if self.sort_by_sessions.get() != "none":
            # Sort by session count
            reverse = self.sort_by_sessions.get() == "desc"
            return sorted(file_data_with_counts, 
                         key=lambda x: x['session_count'], 
                         reverse=reverse)
        else:
            # Sort by filename
            reverse = self.sort_by_name.get() == "desc"
            return sorted(file_data_with_counts, 
                         key=lambda x: x['data']['filename'].lower(), 
                         reverse=reverse)

    def export_to_text_files(self):
        """Export current view to text files with save dialog"""
        try:
            # Choose data to export
            use_filtered = len(self.filtered_file_data) > 0
            data_to_export = self.filtered_file_data if use_filtered and self.filtered_file_data else self.file_data
            
            if not data_to_export:
                messagebox.showwarning("Export Warning", "No data to export")
                return
            
            # Ask user for save location (choose directory)
            selected_dir = filedialog.askdirectory(
                title="Select Directory to Save TXT Files",
                initialdir=self.last_save_dir if hasattr(self, 'last_save_dir') else None
            )
            
            if not selected_dir:
                return  # User cancelled
            
            # Create GROUP_CONNECTION subdirectory
            output_dir = os.path.join(selected_dir, 'GROUP_CONNECTION')
            os.makedirs(output_dir, exist_ok=True)
            
            # Remember the directory for next time
            self.last_save_dir = selected_dir
            
            # Create progress window
            progress_window = ProgressWindow(self.root, "Exporting to TXT")
            
            # Start export in background thread
            export_thread = threading.Thread(
                target=self._export_to_text_files_thread,
                args=(data_to_export, progress_window, output_dir),
                daemon=True
            )
            export_thread.start()
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Error starting TXT export: {e}")
    
    def _export_to_text_files_thread(self, data_to_export, progress_window, output_dir):
        """Export to text files in background thread"""
        try:
            os.makedirs(output_dir, exist_ok=True)
            
            exported_files = []
            total_files = len(data_to_export)
            files_processed = 0
            
            # Export each file with data
            for idx, data in enumerate(data_to_export):
                files_processed += 1
                
                # Update progress with all required arguments
                progress_window.update_progress(
                    f"Processing: {data['filename']}",
                    (idx + 1) / total_files,
                    files_processed,
                    total_files
                )
                
                if not data['sessions']:
                    continue
                
                # Create filename using frequency (e.g., 421-300-000.txt)
                frequency = data.get('frequency', 'unknown')
                output_filename = f"{frequency}.txt"
                output_path = os.path.join(output_dir, output_filename)
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    # Write header
                    f.write(f"File: {data['filename']}\n")
                    f.write(f"Frequency: {data['frequency']}\n")
                    f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("=" * 80 + "\n\n")
                    
                    # Write sessions
                    for group in data['sessions']:
                        if not group:
                            continue
                        
                        # Group header
                        timestamp_str = self.format_output_timestamp(group[0]['timestamp']) if group[0]['timestamp'] else "Unknown time"
                        group_header = f"{timestamp_str} (Sessions: {len(group)})"
                        f.write(group_header + "\n")
                        
                        # Sessions in group
                        for session in group:
                            # Format session line
                            line_parts = []
                            
                            if session['timeslot']:
                                line_parts.append(session['timeslot'])
                            
                            if session['color_code']:
                                line_parts.append(f"CC{session['color_code']}")
                            
                            if session['from'] and session['to']:
                                line_parts.append(f"{session['from']} ─▶ {session['to']}")
                            
                            if session['has_duration']:
                                line_parts.append(f"({session['duration_sec']:.1f}s)")
                            else:
                                line_parts.append("(0.0s)")
                            
                            if session['event']:
                                line_parts.append(session['event'])
                            
                            if session['algorithm'] and session['algorithm'].strip():
                                line_parts.append(f"Alg: {session['algorithm']}")
                                if session['key'] and session['key'].strip():
                                    line_parts.append(f"Key: {session['key']}")
                            
                            session_line = " ".join(line_parts)
                            f.write(session_line + "\n")
                        
                        f.write("\n")  # Empty line between groups
                
                exported_files.append(output_filename)
            
            # Close progress window
            progress_window.close()
            
            # Show success message with only first 3 files
            if exported_files:
                # Show maximum 3 files in the message
                files_preview = "\n".join(exported_files[:3])
                if len(exported_files) > 3:
                    files_preview += f"\n... и ещё {len(exported_files) - 3} файлов"
                
                messagebox.showinfo("Export Complete", 
                                  f"Exported {len(exported_files)} files to:\n{output_dir}\n\n{files_preview}")
            else:
                messagebox.showwarning("Export Warning", "No files with data to export")
                
        except Exception as e:
            progress_window.close()
            messagebox.showerror("Export Error", f"Error creating text files: {e}")

    def export_to_excel(self):
        """Export current view to Excel file with progress window"""
        try:
            # Check if openpyxl is available
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                messagebox.showerror("Export Error", 
                                   "openpyxl library is not installed.\n"
                                   "Install it with: pip install openpyxl")
                return
            
            # Choose data to export
            use_filtered = len(self.filtered_file_data) > 0
            data_to_export = self.filtered_file_data if use_filtered and self.filtered_file_data else self.file_data
            
            # Filter by selected files if any are selected
            if self.selected_files:
                data_to_export = [data for idx, data in enumerate(data_to_export) if idx in self.selected_files]
            
            if not data_to_export:
                messagebox.showwarning("Export Warning", "No data to export")
                return
            
            # Ask user for save location
            filename = f"DMR_Sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = filedialog.asksaveasfilename(
                title="Save Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialdir=self.last_save_dir,
                initialfile=filename
            )
            
            if not file_path:
                return  # User cancelled
            
            # Remember the directory for next time
            self.last_save_dir = os.path.dirname(file_path)
            
            # Create progress window
            progress_window = ProgressWindow(self.root, "Exporting to Excel")
            
            # Start export in background thread
            export_thread = threading.Thread(
                target=self._export_to_excel_thread,
                args=(data_to_export, progress_window, file_path),
                daemon=True
            )
            export_thread.start()
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Error starting Excel export: {e}")
    
    def _export_to_excel_thread(self, data_to_export, progress_window, file_path):
        """Export to Excel in background thread"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DMR Sessions"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Set column headers
            headers = ["File", "Frequency", "Date/Time", "Timeslot", "Color Code", 
                      "From", "To", "Duration (s)", "Event", "Algorithm", "Key"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Set column widths
            column_widths = [20, 15, 20, 10, 12, 12, 12, 12, 20, 15, 15]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Calculate total sessions for progress
            total_sessions = sum(sum(len(group) for group in data['sessions']) for data in data_to_export)
            processed_sessions = 0
            files_processed = 0
            
            # Fill data
            row = 2
            for data_idx, data in enumerate(data_to_export):
                if progress_window.cancelled:
                    break
                
                # Update progress - current file
                progress_window.update_progress(
                    data['filename'], 
                    0, 
                    files_processed, 
                    len(data_to_export)
                )
                
                if not data['sessions']:
                    files_processed += 1
                    continue
                
                for group in data['sessions']:
                    if progress_window.cancelled:
                        break
                    
                    if not group:
                        continue
                    
                    # Group header row
                    timestamp_str = self.format_output_timestamp(group[0]['timestamp']) if group[0]['timestamp'] else "Unknown time"
                    group_header = f"{timestamp_str} (Sessions: {len(group)})"
                    
                    # Merge cells for group header
                    ws.merge_cells(f'A{row}:K{row}')
                    group_cell = ws.cell(row=row, column=1, value=group_header)
                    group_cell.font = Font(bold=True, size=12)
                    group_cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                    group_cell.border = border
                    row += 1
                    
                    # Individual sessions
                    for session in group:
                        if progress_window.cancelled:
                            break
                        
                        ws.cell(row=row, column=1, value=data['filename']).border = border
                        ws.cell(row=row, column=2, value=data['frequency']).border = border
                        ws.cell(row=row, column=3, value=timestamp_str).border = border
                        ws.cell(row=row, column=4, value=session.get('timeslot', '')).border = border
                        ws.cell(row=row, column=5, value=session.get('color_code', '')).border = border
                        ws.cell(row=row, column=6, value=session.get('from', '')).border = border
                        ws.cell(row=row, column=7, value=session.get('to', '')).border = border
                        duration = session.get('duration_sec', 0) if session.get('has_duration', False) else 0
                        ws.cell(row=row, column=8, value=f"{duration:.1f}").border = border
                        ws.cell(row=row, column=9, value=session.get('event', '')).border = border
                        ws.cell(row=row, column=10, value=session.get('algorithm', '')).border = border
                        ws.cell(row=row, column=11, value=session.get('key', '')).border = border
                        row += 1
                        
                        processed_sessions += 1
                        
                        # Update progress every 10 sessions
                        if processed_sessions % 10 == 0:
                            progress = (processed_sessions / total_sessions) * 100
                            progress_window.update_progress(
                                data['filename'], 
                                progress, 
                                files_processed, 
                                len(data_to_export)
                            )
                    
                    # Empty row after group
                    row += 1
                
                files_processed += 1
                
                # Update progress after each file
                progress = (files_processed / len(data_to_export)) * 100
                progress_window.update_progress(
                    data['filename'], 
                    progress, 
                    files_processed, 
                    len(data_to_export)
                )
            
            if not progress_window.cancelled:
                # Save file
                wb.save(file_path)
                
                # Close progress window and show success
                progress_window.close()
                messagebox.showinfo("Export Complete", 
                                  f"Exported to Excel file:\n{os.path.basename(file_path)}")
            else:
                # Export was cancelled
                progress_window.close()
                messagebox.showinfo("Export Cancelled", "Excel export was cancelled.")
            
        except Exception as e:
            progress_window.close()
            messagebox.showerror("Export Error", f"Error creating Excel file: {e}")

    def export_to_pdf(self):
        """Export current view to PDF file with progress window"""
        try:
            # Choose data to export
            use_filtered = len(self.filtered_file_data) > 0
            data_to_export = self.filtered_file_data if use_filtered and self.filtered_file_data else self.file_data
            
            if not data_to_export:
                messagebox.showwarning("Export Warning", "No data to export")
                return
            
            # Ask user for save location
            filename = f"DMR_Sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = filedialog.asksaveasfilename(
                title="Save PDF File",
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialdir=self.last_save_dir,
                initialfile=filename
            )
            
            if not file_path:
                return  # User cancelled
            
            # Remember the directory for next time
            self.last_save_dir = os.path.dirname(file_path)
            
            # Create progress window
            progress_window = ProgressWindow(self.root, "Exporting to PDF")
            
            # Start export in background thread
            export_thread = threading.Thread(
                target=self._export_to_pdf_thread,
                args=(data_to_export, progress_window, file_path),
                daemon=True
            )
            export_thread.start()
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Error starting PDF export: {e}")
    
    def _export_to_pdf_thread(self, data_to_export, progress_window, file_path):
        """Export to PDF in background thread"""
        try:
            # Import ReportLab components
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            
            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=50,
                leftMargin=50,
                topMargin=50,
                bottomMargin=30,
            )
            
            elements = []
            
            # Title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=16,
                spaceAfter=30,
                textColor=colors.HexColor('#000080')
            )
            
            title = Paragraph("DMR Data Report", title_style)
            elements.append(title)
            
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            subtitle = Paragraph(f"Generated: {date_str}", styles['Normal'])
            elements.append(subtitle)
            elements.append(Spacer(1, 0.2*inch))
            
            # Calculate total sessions for progress
            total_sessions = sum(sum(len(group) for group in data['sessions']) for data in data_to_export)
            processed_sessions = 0
            files_processed = 0
            
            # Export files with data
            for data_idx, data in enumerate(data_to_export):
                if progress_window.cancelled:
                    break
                
                # Update progress - current file
                progress_window.update_progress(
                    data['filename'], 
                    0, 
                    files_processed, 
                    len(data_to_export)
                )
                
                if not data['sessions']:
                    files_processed += 1
                    continue
                    
                elements.append(PageBreak())
                elements.append(Paragraph(f"File: {data['filename']}", styles['Heading2']))
                elements.append(Paragraph(f"Frequency: {data['frequency']}", styles['Normal']))
                elements.append(Spacer(1, 0.1*inch))
                
                # Export sessions
                for group in data['sessions']:
                    if progress_window.cancelled:
                        break
                    
                    if not group:
                        continue
                    
                    # Group header
                    timestamp_str = self.format_output_timestamp(group[0]['timestamp']) if group[0]['timestamp'] else "Unknown time"
                    group_header = f"{timestamp_str} (Sessions: {len(group)})"
                    elements.append(Paragraph(group_header, styles['Heading3']))
                    
                    # Sessions in group
                    for session in group:
                        if progress_window.cancelled:
                            break
                        
                        # Format session line
                        line_parts = []
                        
                        if session['timeslot']:
                            line_parts.append(session['timeslot'])
                        
                        if session['color_code']:
                            line_parts.append(f"CC{session['color_code']}")
                        
                        if session['from'] and session['to']:
                            line_parts.append(f"{session['from']} --> {session['to']}")
                        
                        if session['has_duration']:
                            line_parts.append(f"({session['duration_sec']:.1f}s)")
                        else:
                            line_parts.append("(0.0s)")
                        
                        if session['event']:
                            line_parts.append(session['event'])
                        
                        if session['algorithm'] and session['algorithm'].strip():
                            line_parts.append(f"Alg: {session['algorithm']}")
                            if session['key'] and session['key'].strip():
                                line_parts.append(f"Key: {session['key']}")
                        
                        session_line = " ".join(line_parts)
                        elements.append(Paragraph(session_line, styles['Normal']))
                        
                        processed_sessions += 1
                        
                        # Update progress every 10 sessions
                        if processed_sessions % 10 == 0:
                            progress = (processed_sessions / total_sessions) * 100
                            progress_window.update_progress(
                                data['filename'], 
                                progress, 
                                files_processed, 
                                len(data_to_export)
                            )
                    
                    elements.append(Spacer(1, 0.1*inch))
                
                files_processed += 1
                
                # Update progress after each file
                progress = (files_processed / len(data_to_export)) * 100
                progress_window.update_progress(
                    data['filename'], 
                    progress, 
                    files_processed, 
                    len(data_to_export)
                )
            
            if not progress_window.cancelled:
                # Build PDF
                doc.build(elements)
                
                # Close progress window and show success
                progress_window.close()
                messagebox.showinfo("Export Complete", f"PDF report saved to:\n{os.path.basename(file_path)}")
            else:
                # Export was cancelled
                progress_window.close()
                messagebox.showinfo("Export Cancelled", "PDF export was cancelled.")
            
        except Exception as e:
            progress_window.close()
            messagebox.showerror("Export Error", f"Error creating PDF: {e}")


def main():
    root = tk.Tk()
    app = DMRDataViewer(root)
    root.mainloop()


if __name__ == "__main__":
    main()
